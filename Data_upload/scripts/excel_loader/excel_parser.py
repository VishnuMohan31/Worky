"""
Excel Parser Component

This module provides functionality to read and parse Excel files using openpyxl.
It extracts data from sheets and returns structured data as lists of dictionaries.

Requirements: 1.3, 3.1
"""

import logging
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from logging_utils import get_logger, log_info, log_warning, log_error

logger = get_logger("excel_parser")


class ExcelParser:
    """
    Parser for Excel files that extracts sheet data as structured dictionaries.
    
    Handles:
    - Loading Excel workbooks using openpyxl
    - Extracting rows from sheets as list of dictionaries
    - Gracefully handling missing sheets
    - Error handling for corrupted or invalid files
    """
    
    def __init__(self, file_path: str):
        """
        Initialize the Excel parser and load the workbook.
        
        Args:
            file_path: Path to the Excel file (.xlsx)
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file is not a valid Excel file
            Exception: For corrupted or unreadable files
        """
        self.file_path = file_path
        self.workbook: Optional[Workbook] = None
        
        try:
            log_info(logger, 
                f"Loading Excel file: {file_path}",
                file_path=file_path
            )
            
            # Validate file path
            if not file_path:
                raise ValueError("File path cannot be empty")
            
            # Load workbook with data_only=True to get calculated values instead of formulas
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            log_info(logger, 
                f"Successfully loaded workbook with {len(self.workbook.sheetnames)} sheets",
                file_path=file_path,
                sheet_count=len(self.workbook.sheetnames),
                sheet_names=self.workbook.sheetnames
            )
            
        except FileNotFoundError as e:
            error_msg = f"Excel file not found: {file_path}"
            log_error(logger, error_msg, exc_info=True)
            raise FileNotFoundError(error_msg) from e
            
        except openpyxl.utils.exceptions.InvalidFileException as e:
            error_msg = f"Invalid Excel file format: {file_path}. Please ensure the file is a valid .xlsx or .xls file."
            log_error(logger, error_msg, exc_info=True)
            raise ValueError(error_msg) from e
            
        except PermissionError as e:
            error_msg = f"Permission denied accessing file: {file_path}"
            log_error(logger, error_msg, exc_info=True)
            raise PermissionError(error_msg) from e
            
        except Exception as e:
            error_msg = f"Failed to load Excel file: {file_path}. Error: {str(e)}"
            log_error(logger, error_msg, exc_info=True)
            raise Exception(f"Corrupted or unreadable Excel file: {file_path}") from e

    def get_sheet_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Extract all rows from a sheet as a list of dictionaries.
        
        The first row is treated as headers, and subsequent rows are converted
        to dictionaries with header names as keys.
        
        Args:
            sheet_name: Name of the sheet to extract data from
            
        Returns:
            List of dictionaries where each dict represents a row.
            Returns empty list if sheet doesn't exist or has no data.
            
        Example:
            If sheet has:
            | Name  | Age | City     |
            | Alice | 30  | New York |
            | Bob   | 25  | Boston   |
            
            Returns:
            [
                {'Name': 'Alice', 'Age': 30, 'City': 'New York'},
                {'Name': 'Bob', 'Age': 25, 'City': 'Boston'}
            ]
        """
        if not self.workbook:
            log_warning(logger, "Workbook not loaded")
            return []
        
        if not sheet_name:
            log_warning(logger, "Sheet name cannot be empty")
            return []
        
        # Check if sheet exists
        if sheet_name not in self.workbook.sheetnames:
            log_warning(logger, f"Sheet '{sheet_name}' not found in workbook. Available sheets: {self.workbook.sheetnames}")
            return []
        
        try:
            sheet: Worksheet = self.workbook[sheet_name]
            
            # Check if sheet has any data
            if sheet.max_row < 2:  # Need at least header + 1 data row
                log_info(logger, f"Sheet '{sheet_name}' is empty or has only headers")
                return []
            
            # Extract headers from first row
            headers = []
            for cell in sheet[1]:
                header_value = cell.value
                if header_value is not None:
                    headers.append(str(header_value).strip())
                else:
                    headers.append(f"Column_{cell.column}")
            
            if not headers:
                log_warning(logger, f"Sheet '{sheet_name}' has no headers")
                return []
            
            log_info(logger, 
                f"Sheet '{sheet_name}': Found {len(headers)} columns, {sheet.max_row - 1} data rows",
                sheet_name=sheet_name,
                column_count=len(headers),
                data_row_count=sheet.max_row - 1,
                headers=headers
            )
            
            # Extract data rows
            data_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip completely empty rows
                    if all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
                        continue
                    
                    # Create dictionary for this row
                    row_dict = {}
                    for col_idx, (header, value) in enumerate(zip(headers, row)):
                        # Store value as-is, let DataMapper handle conversions
                        row_dict[header] = value
                    
                    # Add row number for debugging purposes
                    row_dict['_row_number'] = row_idx
                    data_rows.append(row_dict)
                    
                except Exception as row_error:
                    log_error(logger, f"Error processing row {row_idx} in sheet '{sheet_name}': {str(row_error)}")
                    # Continue processing other rows
                    continue
            
            log_info(logger, 
                f"Sheet '{sheet_name}': Extracted {len(data_rows)} non-empty rows",
                sheet_name=sheet_name,
                extracted_rows=len(data_rows),
                total_rows=sheet.max_row - 1
            )
            return data_rows
            
        except KeyError as e:
            log_error(logger, f"Sheet access error for '{sheet_name}': {str(e)}", exc_info=True)
            return []
        except Exception as e:
            log_error(logger, f"Unexpected error reading sheet '{sheet_name}': {str(e)}", exc_info=True)
            return []

    def get_available_sheets(self) -> List[str]:
        """
        Return list of all sheet names in the workbook.
        
        Returns:
            List of sheet names, or empty list if workbook not loaded
        """
        if not self.workbook:
            log_warning(logger, "Workbook not loaded")
            return []
        
        sheet_names = self.workbook.sheetnames
        log_info(logger, f"Available sheets: {sheet_names}")
        return sheet_names
    
    def close(self):
        """
        Close the workbook and clean up resources.
        
        Should be called when done processing the Excel file to free memory.
        """
        if self.workbook:
            try:
                self.workbook.close()
                log_info(logger, f"Closed workbook: {self.file_path}")
                self.workbook = None
            except Exception as e:
                log_error(logger, f"Error closing workbook: {str(e)}", exc_info=True)
                # Still set to None to prevent further use
                self.workbook = None
        else:
            log_info(logger, "No workbook to close")
    
    def __enter__(self):
        """Context manager entry - returns self for use in 'with' statements."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - ensures workbook is closed."""
        self.close()
        return False  # Don't suppress exceptions
